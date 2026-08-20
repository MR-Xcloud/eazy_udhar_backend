"""Restore a seller's data from the EOD Excel backup (see eod_excel_report.py for the
format this reads back).

Additive only: existing customers and transactions are never modified or deleted.
Customers are matched by phone; transactions have no stable ID in the export, so a
duplicate is anything matching an existing row's customer, type, amount, note, and
effective minute — meaning the same backup file can be re-uploaded safely.

A restored customer's outstanding/advance balances are updated by applying just the
newly-inserted rows on top of whatever the customer's balance currently is, using the
same running-balance rules `services.add_credit` / `receive_payment` / `deposit_advance`
/ `use_advance` apply — not by re-deriving the total from the full transaction history.
That's deliberate: on this production data, a customer's live outstanding_amount does
not always equal the sum of their LedgerTransaction rows (manual corrections and other
off-ledger adjustments happen over an account's life), so a full replay would silently
discard those adjustments for anyone who already has history. Starting from the current
balance and only applying the restored delta preserves them. It's also why those
side-effecting service functions aren't called directly — re-running their business
logic (wallet auto-apply, SMS, notifications) would double-apply history already
reflected as separate rows in the export, or spam customers about events long past.
"""
from datetime import datetime, time as dt_time, timedelta
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.db.models.functions import Coalesce
from django.utils import timezone
from openpyxl import load_workbook

from customerapp.messaging import link_seller_customer
from easyudhar.payment_utils import normalize_payment_method

from .eod_excel_report import TYPE_LABELS
from .models import LedgerTransaction, SellerCustomer
from .services import refresh_due_status
from .sync_service import find_customer_by_phone
from .utils import normalize_phone

TYPE_LABELS_REVERSE = {label: key for key, label in TYPE_LABELS.items()}

REQUIRED_SHEETS = ('Summary', 'Customers', 'Transactions')
CUSTOMERS_HEADERS = ['Customer', 'Phone', 'Outstanding (Rs.)', 'Status', 'Advance Deposited (Rs.)', 'Next Due Date']
TRANSACTIONS_HEADERS = [
    'Date', 'Time', 'Customer', 'Phone', 'Type',
    'Amount (Rs.)', 'Payment Method', 'Note', 'Due Date', 'Recorded At',
]


class RestoreError(Exception):
    def __init__(self, message, *, code='bad_format'):
        super().__init__(message)
        self.message = message
        self.code = code


def _find_summary_field(ws, label):
    for row in range(1, 20):
        if str(ws.cell(row=row, column=1).value or '').strip() == label:
            return ws.cell(row=row, column=2).value
    return None


def _read_headers(ws):
    return [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]


def _parse_date_only(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'year'):
        return value
    from django.utils.dateparse import parse_date

    return parse_date(str(value))


def _parse_effective_datetime(date_val, time_val):
    day = _parse_date_only(date_val)
    if day is None:
        return None
    hour, minute = 0, 0
    if time_val:
        try:
            parts = str(time_val).strip().split(':')
            hour, minute = int(parts[0]), int(parts[1])
        except (ValueError, IndexError):
            hour, minute = 0, 0
    naive = datetime.combine(day, dt_time(hour=hour, minute=minute))
    return timezone.make_aware(naive, timezone.get_current_timezone())


def _decimal(value):
    if value is None or value == '':
        return Decimal('0')
    try:
        return Decimal(str(value)).quantize(Decimal('0.01'))
    except InvalidOperation:
        return Decimal('0')


def parse_backup_workbook(fileobj):
    try:
        wb = load_workbook(fileobj, data_only=True, read_only=True)
    except Exception as exc:
        raise RestoreError(f'Could not read this file as an Excel workbook: {exc}', code='unreadable') from exc

    try:
        missing = [name for name in REQUIRED_SHEETS if name not in wb.sheetnames]
        if missing:
            raise RestoreError(
                f"This doesn't look like an EazyUdhar backup — missing sheet(s): {', '.join(missing)}.",
            )

        summary_ws = wb['Summary']
        business = str(_find_summary_field(summary_ws, 'Business') or '').strip()
        email = str(_find_summary_field(summary_ws, 'Seller Email') or '').strip()

        customers_ws = wb['Customers']
        if _read_headers(customers_ws)[:len(CUSTOMERS_HEADERS)] != CUSTOMERS_HEADERS:
            raise RestoreError('The Customers sheet columns do not match the EazyUdhar backup format.')

        transactions_ws = wb['Transactions']
        if _read_headers(transactions_ws)[:len(TRANSACTIONS_HEADERS)] != TRANSACTIONS_HEADERS:
            raise RestoreError('The Transactions sheet columns do not match the EazyUdhar backup format.')

        customer_rows = []
        for row in customers_ws.iter_rows(min_row=2, values_only=True):
            if not row or all(v in (None, '') for v in row):
                continue
            name, phone, outstanding, _status_label, advance_deposited, next_due_date = row[:6]
            if not phone:
                continue
            customer_rows.append({
                'name': str(name or '').strip() or 'Customer',
                'phone': str(phone).strip(),
                'outstanding': outstanding,
                'advance_deposited': advance_deposited,
                'next_due_date': _parse_date_only(next_due_date),
            })

        transaction_rows = []
        row_num = 1
        for row in transactions_ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if not row or all(v in (None, '') for v in row):
                continue
            date_val, time_val, cust_name, phone, type_label, amount, method_label, note, due_date, _recorded_at = row[:10]
            transaction_rows.append({
                'row': row_num,
                'effective_at': _parse_effective_datetime(date_val, time_val),
                'customer_name': str(cust_name or '').strip(),
                'phone': str(phone or '').strip(),
                'type_label': str(type_label or '').strip(),
                'amount': _decimal(amount),
                'payment_method_label': str(method_label or '').strip(),
                'note': str(note or '').strip(),
                'due_date': _parse_date_only(due_date),
            })
    finally:
        wb.close()

    return {
        'business': business,
        'email': email,
        'customer_rows': customer_rows,
        'transaction_rows': transaction_rows,
    }


def _is_duplicate_transaction(customer, tx_type, amount, effective_at, note, *, before):
    """Duplicate check is only against rows that existed before this import run —
    never against rows this same run has already inserted. Two genuinely distinct
    transactions in the source file can share type/amount/note/minute (e.g. two
    separate cash payments moments apart with no note), and comparing against our
    own inserts would wrongly collapse them into one."""
    qs = LedgerTransaction.objects.filter(
        customer=customer, transaction_type=tx_type, amount=amount, created_at__lt=before,
    )
    qs = qs.filter(note=note) if note else qs.filter(note='')
    if effective_at is not None:
        start = effective_at.replace(second=0, microsecond=0)
        qs = qs.annotate(eff=Coalesce('device_created_at', 'created_at')).filter(
            eff__gte=start, eff__lt=start + timedelta(minutes=1)
        )
    return qs.exists()


def _apply_new_transactions(customer, new_txs, *, update_due_date):
    """Apply just the newly-restored rows (in the order they were inserted, i.e. the
    file's chronological order) on top of the customer's current balance — never a
    full re-derivation from history. See module docstring for why.

    `update_due_date` is only true for brand-new customers: for an existing customer,
    a restored historical credit shouldn't silently move their current reminder due
    date around, so their next_due_date is left untouched by the merge.
    """
    outstanding = customer.outstanding_amount
    advance_deposited = customer.advance_deposited
    advance_used = customer.advance_used
    next_due_date = customer.next_due_date

    for tx in new_txs:
        if tx.transaction_type == LedgerTransaction.TYPE_CREDIT:
            outstanding += tx.amount
            if update_due_date and tx.due_date:
                next_due_date = tx.due_date
        elif tx.transaction_type == LedgerTransaction.TYPE_PAYMENT:
            applied = min(tx.amount, outstanding)
            outstanding -= applied
            advance_deposited += (tx.amount - applied)
        elif tx.transaction_type == LedgerTransaction.TYPE_ADVANCE_DEPOSIT:
            advance_deposited += tx.amount
        elif tx.transaction_type == LedgerTransaction.TYPE_ADVANCE_USE:
            advance_used += tx.amount

    customer.outstanding_amount = outstanding
    customer.advance_deposited = advance_deposited
    customer.advance_used = advance_used
    customer.next_due_date = next_due_date
    customer.save(update_fields=[
        'outstanding_amount', 'advance_deposited', 'advance_used', 'next_due_date', 'updated_at',
    ])
    refresh_due_status(customer)


def _create_customer(seller, *, name, phone, next_due_date=None):
    customer = SellerCustomer(seller=seller, name=name or 'Customer', phone=phone, next_due_date=next_due_date)
    customer.sync_composed_address()
    customer.save()
    return customer


@transaction.atomic
def restore_seller_backup(seller, fileobj, *, force_identity_mismatch=False):
    parsed = parse_backup_workbook(fileobj)

    identity = {
        'file_business': parsed['business'],
        'file_email': parsed['email'],
        'seller_business': seller.business_name or '',
        'seller_email': seller.email or '',
        'matched': True,
    }
    if not parsed['email']:
        identity['matched'] = False
        if not force_identity_mismatch:
            raise RestoreError(
                "Could not find a seller email in this file's Summary sheet, so it can't be "
                "verified as this seller's backup.",
                code='identity_unverifiable',
            )
    elif parsed['email'].strip().lower() != (seller.email or '').strip().lower():
        identity['matched'] = False
        if not force_identity_mismatch:
            raise RestoreError(
                f"This backup belongs to {parsed['email']} ({parsed['business'] or 'unknown business'}), "
                f"not {seller.email}. Confirm this is the right seller before restoring.",
                code='identity_mismatch',
            )

    import_started_at = timezone.now()
    warnings = []
    customers_created = 0
    customers_matched = 0
    newly_created_ids = set()
    phone_to_customer = {}

    for row in parsed['customer_rows']:
        norm_phone = normalize_phone(row['phone'])
        if not norm_phone:
            warnings.append(f"Customers sheet: '{row['name']}' has no usable phone number, skipped.")
            continue
        existing = find_customer_by_phone(seller, row['phone'])
        if existing:
            customers_matched += 1
            phone_to_customer[norm_phone] = existing
            continue
        customer = _create_customer(seller, name=row['name'], phone=row['phone'], next_due_date=row['next_due_date'])
        customers_created += 1
        newly_created_ids.add(customer.id)
        phone_to_customer[norm_phone] = customer

    transactions_imported = 0
    transactions_skipped_duplicate = 0
    transactions_skipped_error = 0
    new_txs_by_customer_id = {}

    for row in parsed['transaction_rows']:
        tx_type = TYPE_LABELS_REVERSE.get(row['type_label'])
        if not tx_type:
            warnings.append(f"Transactions row {row['row']}: unrecognized type '{row['type_label']}', skipped.")
            transactions_skipped_error += 1
            continue

        norm_phone = normalize_phone(row['phone'])
        if not norm_phone:
            warnings.append(f"Transactions row {row['row']}: missing/invalid phone number, skipped.")
            transactions_skipped_error += 1
            continue

        customer = phone_to_customer.get(norm_phone)
        if customer is None:
            customer = find_customer_by_phone(seller, row['phone'])
            if customer is None:
                customer = _create_customer(seller, name=row['customer_name'], phone=row['phone'])
                customers_created += 1
                newly_created_ids.add(customer.id)
                warnings.append(
                    f"Transactions row {row['row']}: customer '{row['customer_name']}' ({row['phone']}) "
                    "wasn't in the Customers sheet — created from this transaction row."
                )
            phone_to_customer[norm_phone] = customer

        if row['amount'] <= 0:
            warnings.append(f"Transactions row {row['row']}: non-positive amount, skipped.")
            transactions_skipped_error += 1
            continue

        if _is_duplicate_transaction(
            customer, tx_type, row['amount'], row['effective_at'], row['note'], before=import_started_at,
        ):
            transactions_skipped_duplicate += 1
            continue

        tx = LedgerTransaction.objects.create(
            seller=seller,
            customer=customer,
            transaction_type=tx_type,
            amount=row['amount'],
            note=row['note'],
            payment_method=normalize_payment_method(row['payment_method_label']) if row['payment_method_label'] else '',
            due_date=row['due_date'] if tx_type == LedgerTransaction.TYPE_CREDIT else None,
            device_created_at=row['effective_at'],
        )
        transactions_imported += 1
        new_txs_by_customer_id.setdefault(customer.id, []).append(tx)

    for customer_id, new_txs in new_txs_by_customer_id.items():
        customer = SellerCustomer.objects.select_for_update().get(pk=customer_id)
        _apply_new_transactions(customer, new_txs, update_due_date=customer_id in newly_created_ids)
        link_seller_customer(customer)

    # Brand-new customers that ended up with zero transactions (present only in the
    # Customers sheet) have nothing to apply — take the sheet's snapshot directly.
    for row in parsed['customer_rows']:
        norm_phone = normalize_phone(row['phone'])
        customer = phone_to_customer.get(norm_phone)
        if customer and customer.id in newly_created_ids and customer.id not in new_txs_by_customer_id:
            customer.outstanding_amount = _decimal(row['outstanding'])
            customer.advance_deposited = _decimal(row['advance_deposited'])
            customer.save(update_fields=['outstanding_amount', 'advance_deposited', 'updated_at'])
            refresh_due_status(customer)
            link_seller_customer(customer)

    return {
        'identity': identity,
        'customers_created': customers_created,
        'customers_matched': customers_matched,
        'customers_balance_updated': len(new_txs_by_customer_id),
        'transactions_imported': transactions_imported,
        'transactions_skipped_duplicate': transactions_skipped_duplicate,
        'transactions_skipped_error': transactions_skipped_error,
        'warnings': warnings,
    }
