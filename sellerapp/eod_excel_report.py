"""End-of-day Excel transaction backup, emailed to each seller for their own records.

Every query here is filtered by a single seller, so a seller only ever receives
their own ledger data — never another seller's transactions or customers.
"""
import io
from collections import OrderedDict
from decimal import Decimal

from django.conf import settings
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, Sum
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from easyudhar.payment_utils import payment_method_label

from .models import LedgerTransaction, Seller, SellerCustomer, SellerSettings
from .daily_summary import _parse_hhmm, build_seller_summary

HEADER_FILL = PatternFill('solid', fgColor='0F766E')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FILL = PatternFill('solid', fgColor='0F766E')
TITLE_FONT = Font(bold=True, size=17, color='FFFFFF')
SECTION_FILL = PatternFill('solid', fgColor='14B8A6')
SECTION_FONT = Font(bold=True, size=11, color='FFFFFF')
LABEL_FONT = Font(bold=True, color='334155')
ZEBRA_FILL = PatternFill('solid', fgColor='F8FAFC')
THIN_SIDE = Side(style='thin', color='D1D5DB')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CURRENCY_FMT = '"Rs. "#,##0.00'
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

TYPE_LABELS = {
    LedgerTransaction.TYPE_CREDIT: 'Credit Given',
    LedgerTransaction.TYPE_PAYMENT: 'Payment Received',
    LedgerTransaction.TYPE_ADVANCE_DEPOSIT: 'Advance Deposit',
    LedgerTransaction.TYPE_ADVANCE_USE: 'Advance Used',
}

# Soft fills + matching bold fonts, keyed by transaction type — used to color-code rows
# in the Transactions sheet so a seller can scan money-in vs money-out at a glance.
TYPE_COLORS = {
    LedgerTransaction.TYPE_CREDIT: (PatternFill('solid', fgColor='FEF2F2'), Font(color='B91C1C', bold=True)),
    LedgerTransaction.TYPE_PAYMENT: (PatternFill('solid', fgColor='F0FDF4'), Font(color='15803D', bold=True)),
    LedgerTransaction.TYPE_ADVANCE_DEPOSIT: (PatternFill('solid', fgColor='EFF6FF'), Font(color='1D4ED8', bold=True)),
    LedgerTransaction.TYPE_ADVANCE_USE: (PatternFill('solid', fgColor='FFF7ED'), Font(color='C2410C', bold=True)),
}

STATUS_LABELS = {
    SellerCustomer.STATUS_PENDING: 'Pending',
    SellerCustomer.STATUS_OVERDUE: 'Overdue',
    SellerCustomer.STATUS_PAID: 'Paid',
    SellerCustomer.STATUS_SETTLED: 'Settled',
}
STATUS_FONTS = {
    SellerCustomer.STATUS_OVERDUE: Font(color='B91C1C', bold=True),
    SellerCustomer.STATUS_PENDING: Font(color='C2410C', bold=True),
    SellerCustomer.STATUS_PAID: Font(color='15803D', bold=True),
    SellerCustomer.STATUS_SETTLED: Font(color='1D4ED8', bold=True),
}


def is_eod_backup_due_now(settings_row, now=None):
    now = now or timezone.localtime()
    hour, minute = _parse_hhmm(settings_row.eod_excel_backup_time)
    return now.hour == hour and now.minute < 15


def _style_header_row(ws, row=1, last_col=1, start_col=1):
    for col in range(start_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def _set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _section_band(ws, row, text, start_col, last_col):
    """A filled, merged banner used to introduce a table section (distinct from
    the darker column-header row that follows it)."""
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=start_col, value=text)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for col in range(start_col, last_col + 1):
        ws.cell(row=row, column=col).fill = SECTION_FILL
    ws.row_dimensions[row].height = 20


def _seller_totals(seller):
    customers = SellerCustomer.objects.filter(seller=seller)
    outstanding = customers.aggregate(t=Sum('outstanding_amount'))['t'] or Decimal('0')
    overdue_qs = customers.filter(status=SellerCustomer.STATUS_OVERDUE)
    overdue_count = overdue_qs.count()
    overdue_amount = overdue_qs.aggregate(t=Sum('outstanding_amount'))['t'] or Decimal('0')
    total_credit = (
        LedgerTransaction.objects.filter(
            seller=seller, transaction_type=LedgerTransaction.TYPE_CREDIT
        ).aggregate(t=Sum('amount'))['t']
        or Decimal('0')
    )
    total_received = (
        LedgerTransaction.objects.filter(
            seller=seller, transaction_type=LedgerTransaction.TYPE_PAYMENT
        ).aggregate(t=Sum('amount'))['t']
        or Decimal('0')
    )
    return {
        'customers_count': customers.count(),
        'outstanding': outstanding,
        'overdue_count': overdue_count,
        'overdue_amount': overdue_amount,
        'total_credit': total_credit,
        'total_received': total_received,
    }


def _write_kv_rows(ws, start_row, rows, currency_labels=()):
    row_idx = start_row
    for label, value in rows:
        label_cell = ws.cell(row=row_idx, column=1, value=label)
        label_cell.font = LABEL_FONT
        label_cell.alignment = LEFT_ALIGN
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.alignment = RIGHT_ALIGN
        if isinstance(value, float):
            value_cell.number_format = CURRENCY_FMT if label in currency_labels else '#,##0'
        if label:
            fill = ZEBRA_FILL if row_idx % 2 else PatternFill(fill_type=None)
            for cell in (label_cell, value_cell):
                cell.fill = fill
                cell.border = THIN_BORDER
        row_idx += 1
    return row_idx


def _monthly_collections(seller, months=6):
    """Payment totals for the last N months, oldest first.

    Bucketed in Python (not via TruncMonth) because this DB has no MySQL
    timezone tables loaded, which makes server-side tz-aware date truncation
    raise — the same reason build_seller_summary() avoids TruncDate.
    """
    now = timezone.localtime()
    by_month = {}
    for tx in LedgerTransaction.objects.filter(
        seller=seller, transaction_type=LedgerTransaction.TYPE_PAYMENT
    ).only('amount', 'created_at', 'device_created_at'):
        local_effective = timezone.localtime(tx.effective_at)
        key = f'{local_effective.year:04d}-{local_effective.month:02d}'
        by_month[key] = by_month.get(key, Decimal('0')) + tx.amount

    ordered = OrderedDict()
    cursor_year, cursor_month = now.year, now.month
    keys = []
    for _ in range(months):
        keys.append((cursor_year, cursor_month))
        cursor_month -= 1
        if cursor_month == 0:
            cursor_month = 12
            cursor_year -= 1
    for year, month in reversed(keys):
        key = f'{year:04d}-{month:02d}'
        label = timezone.datetime(year, month, 1).strftime('%b %Y')
        ordered[label] = float(by_month.get(key) or 0)
    return ordered


def _type_breakdown(seller):
    qs = (
        LedgerTransaction.objects.filter(seller=seller)
        .values('transaction_type')
        .annotate(count=Count('id'), total=Sum('amount'))
    )
    by_type = {row['transaction_type']: row for row in qs}
    return [
        (TYPE_LABELS[t], by_type.get(t, {}).get('count', 0) or 0, float(by_type.get(t, {}).get('total') or 0))
        for t in TYPE_LABELS
    ]


def _top_customers(seller, limit=20):
    qs = (
        SellerCustomer.objects.filter(seller=seller, outstanding_amount__gt=0)
        .order_by('-outstanding_amount')[:limit]
    )
    return [(c.name, c.phone, float(c.outstanding_amount)) for c in qs]


def _write_table_row(ws, row, values, *, currency_cols=(), border=True):
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        if col in currency_cols:
            cell.number_format = CURRENCY_FMT
            cell.alignment = RIGHT_ALIGN
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
            cell.number_format = '#,##0'
            cell.alignment = RIGHT_ALIGN
        else:
            cell.alignment = LEFT_ALIGN
        if border:
            cell.border = THIN_BORDER


def _build_summary_sheet(ws, seller, generated_at, tx_count, totals):
    """Combined Summary + Insights sheet: overview totals, monthly collections,
    transaction mix, and top customers by outstanding — tables only, no charts.
    """
    ws.sheet_view.showGridLines = False

    last_col = 9  # A..I, matches _set_widths below
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title_cell = ws.cell(row=1, column=1, value='EazyUdhar — End of Day Transaction Backup')
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    for col in range(1, last_col + 1):
        ws.cell(row=1, column=col).fill = TITLE_FILL
    ws.row_dimensions[1].height = 30

    # --- Overview ---
    _section_band(ws, 3, 'Overview', 1, 2)
    overview_rows = [
        ('Business', seller.business_name or seller.full_name or '-'),
        ('Seller Email', seller.email),
        ('Generated At', timezone.localtime(generated_at).strftime('%Y-%m-%d %H:%M %Z')),
        ('Total Customers', totals['customers_count']),
        ('Total Transactions (all time)', tx_count),
        ('Total Credit Given', float(totals['total_credit'])),
        ('Total Collected', float(totals['total_received'])),
        ('Current Outstanding', float(totals['outstanding'])),
        ('Overdue Customers', totals['overdue_count']),
        ('Overdue Amount', float(totals['overdue_amount'])),
    ]
    currency_labels = {'Total Credit Given', 'Total Collected', 'Current Outstanding', 'Overdue Amount'}
    next_row = _write_kv_rows(ws, 4, overview_rows, currency_labels=currency_labels)

    # --- Monthly collections ---
    section_row = next_row + 2
    _section_band(ws, section_row, 'Monthly Collections (last 6 months)', 1, 2)
    header_row = section_row + 1
    ws.cell(row=header_row, column=1, value='Month')
    ws.cell(row=header_row, column=2, value='Collected')
    _style_header_row(ws, row=header_row, last_col=2)

    row = header_row + 1
    for label, total in _monthly_collections(seller).items():
        _write_table_row(ws, row, [label, total], currency_cols={2})
        row += 1

    # --- Transaction mix ---
    section_row = row + 2
    _section_band(ws, section_row, 'Transaction Mix (all time)', 1, 3)
    header_row = section_row + 1
    ws.cell(row=header_row, column=1, value='Type')
    ws.cell(row=header_row, column=2, value='Count')
    ws.cell(row=header_row, column=3, value='Amount')
    _style_header_row(ws, row=header_row, last_col=3)

    row = header_row + 1
    for label, count, total in _type_breakdown(seller):
        _write_table_row(ws, row, [label, count, total], currency_cols={3})
        row += 1

    # --- Top customers by outstanding (placed beside the sections above, starting col G) ---
    top_col = 7  # column G
    _section_band(ws, 3, 'Top 20 Customers by Outstanding', top_col, top_col + 2)
    header_row = 4
    ws.cell(row=header_row, column=top_col, value='Customer')
    ws.cell(row=header_row, column=top_col + 1, value='Phone')
    ws.cell(row=header_row, column=top_col + 2, value='Outstanding')
    _style_header_row(ws, row=header_row, start_col=top_col, last_col=top_col + 2)

    top_row = header_row + 1
    for name, phone, amount in _top_customers(seller):
        for offset, value in enumerate((name, phone, amount)):
            col = top_col + offset
            cell = ws.cell(row=top_row, column=col, value=value)
            cell.border = THIN_BORDER
            if col == top_col + 2:
                cell.number_format = CURRENCY_FMT
                cell.alignment = RIGHT_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
        top_row += 1

    ws.freeze_panes = 'A4'
    _set_widths(ws, [30, 28, 18, 3, 3, 3, 28, 18, 18])


def _build_customers_sheet(ws, seller):
    """Every customer this seller has, ranked by outstanding balance (highest first)."""
    ws.sheet_view.showGridLines = False
    headers = ['Customer', 'Phone', 'Outstanding (Rs.)', 'Status', 'Advance Deposited (Rs.)', 'Next Due Date']
    ws.append(headers)
    _style_header_row(ws, last_col=len(headers))

    customers = SellerCustomer.objects.filter(seller=seller).order_by('-outstanding_amount', 'name')
    for c in customers:
        ws.append([
            c.name,
            c.phone,
            float(c.outstanding_amount),
            STATUS_LABELS.get(c.status, c.status),
            float(c.advance_deposited),
            c.next_due_date,
        ])
        r = ws.max_row
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        if c.next_due_date:
            ws.cell(row=r, column=6).number_format = 'yyyy-mm-dd'
        status_font = STATUS_FONTS.get(c.status)
        if status_font:
            ws.cell(row=r, column=4).font = status_font
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = THIN_BORDER

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}'
    _set_widths(ws, [26, 16, 18, 14, 20, 14])


def build_transactions_workbook(seller):
    """Full ledger snapshot for a single seller: Summary (with insights) + Transactions sheets."""
    generated_at = timezone.now()

    txs = (
        LedgerTransaction.objects.filter(seller=seller)
        .select_related('customer')
        .order_by('created_at')
    )
    tx_count = txs.count()
    totals = _seller_totals(seller)

    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = 'Summary'
    _build_summary_sheet(summary_ws, seller, generated_at, tx_count, totals)

    customers_ws = wb.create_sheet('Customers')
    _build_customers_sheet(customers_ws, seller)

    ws = wb.create_sheet('Transactions')
    headers = [
        'Date', 'Time', 'Customer', 'Phone', 'Type',
        'Amount (Rs.)', 'Payment Method', 'Note', 'Due Date', 'Recorded At',
    ]
    ws.append(headers)
    _style_header_row(ws, last_col=len(headers))

    for tx in txs:
        local_effective = timezone.localtime(tx.effective_at)
        row_values = [
            local_effective.date(),
            local_effective.strftime('%H:%M'),
            tx.customer.name,
            tx.customer.phone,
            TYPE_LABELS.get(tx.transaction_type, tx.transaction_type),
            float(tx.amount),
            payment_method_label(tx.payment_method) if tx.payment_method else '',
            tx.note,
            tx.due_date,
            timezone.localtime(tx.created_at).strftime('%Y-%m-%d %H:%M'),
        ]
        ws.append(row_values)
        r = ws.max_row
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=6).number_format = '#,##0.00'
        if tx.due_date:
            ws.cell(row=r, column=9).number_format = 'yyyy-mm-dd'

        row_fill, type_font = TYPE_COLORS.get(tx.transaction_type, (None, None))
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill
        if type_font:
            ws.cell(row=r, column=5).font = type_font
            ws.cell(row=r, column=6).font = type_font

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(ws.max_row, 1)}'
    _set_widths(ws, [12, 8, 22, 14, 18, 14, 16, 32, 12, 18])

    return wb


def render_workbook_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _format_inr(amount):
    return f'Rs. {float(amount):,.2f}'


def _build_email(seller, *, tx_count, totals, attachment_name):
    today = timezone.localdate()
    display_name = seller.full_name or seller.business_name or 'there'
    summary = build_seller_summary(seller)

    top_overdue = (
        SellerCustomer.objects.filter(seller=seller, status=SellerCustomer.STATUS_OVERDUE)
        .order_by('-outstanding_amount')
        .first()
    )

    context = {
        'display_name': display_name,
        'report_date': today.strftime('%d %b %Y'),
        'today_collection_display': _format_inr(summary['today_collection']),
        'outstanding_display': _format_inr(totals['outstanding']),
        'overdue_count': totals['overdue_count'],
        'overdue_amount_display': _format_inr(totals['overdue_amount']),
        'tx_count': tx_count,
        'top_overdue_name': top_overdue.name if top_overdue else '',
        'top_overdue_amount_display': _format_inr(top_overdue.outstanding_amount) if top_overdue else '',
        'attachment_name': attachment_name,
        'logo_url': getattr(settings, 'OTP_EMAIL_LOGO_URL', ''),
    }
    html_body = render_to_string('emails/eod_backup.html', context)
    text_body = (
        f'Hi {display_name},\n\n'
        f'Your EazyUdhar business snapshot for {context["report_date"]}:\n'
        f'  Collected today: {context["today_collection_display"]}\n'
        f'  Outstanding: {context["outstanding_display"]}\n'
        f'  Overdue: {context["overdue_count"]} customer(s), {context["overdue_amount_display"]}\n'
        f'  Total transactions on file: {tx_count}\n\n'
        f'Attached ({attachment_name}) is your full transaction backup — Summary & Insights, '
        'every Customer ranked by outstanding, and the complete transaction log.\n\n'
        '— EazyUdhar, powered by INWIZY'
    )
    subject = f'EazyUdhar — Daily Transaction Backup ({context["report_date"]})'
    return subject, text_body, html_body


def send_eod_backup_for_seller(seller, *, force=False, to_email=None):
    """Build and send this seller's own EOD Excel backup.

    `to_email` overrides the recipient for manual/admin test sends only — the
    report content itself is always built strictly from this seller's own data.
    """
    if not seller.is_active:
        return {'skipped': True, 'reason': 'seller suspended'}

    from .excel_report_addon_service import has_excel_report_access

    if not force and not has_excel_report_access(seller):
        return {'skipped': True, 'reason': 'addon not active'}

    settings_row, _ = SellerSettings.objects.get_or_create(seller=seller)
    if not settings_row.eod_excel_backup_enabled and not force:
        return {'skipped': True, 'reason': 'disabled'}

    now = timezone.localtime()
    if not force and not is_eod_backup_due_now(settings_row, now):
        return {'skipped': True, 'reason': 'not scheduled time'}

    recipient = to_email or seller.email
    if not recipient:
        return {'skipped': True, 'reason': 'no email on file'}

    txs = LedgerTransaction.objects.filter(seller=seller)
    tx_count = txs.count()
    totals = _seller_totals(seller)

    wb = build_transactions_workbook(seller)
    content = render_workbook_bytes(wb)

    today = timezone.localdate()
    safe_business = (seller.business_name or 'Seller').strip().replace(' ', '_')
    filename = f'EazyUdhar-Transactions-{safe_business}-{today.isoformat()}.xlsx'

    subject, text_body, html_body = _build_email(
        seller, tx_count=tx_count, totals=totals, attachment_name=filename
    )
    email = EmailMultiAlternatives(
        subject=subject,
        body=text_body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[recipient],
    )
    email.attach_alternative(html_body, 'text/html')
    email.attach(
        filename,
        content,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    sent = email.send(fail_silently=False)
    return {'sent': bool(sent), 'filename': filename, 'to': recipient}


def run_eod_backups(*, force=False):
    results = []
    for seller in Seller.objects.filter(is_active=True):
        try:
            row = send_eod_backup_for_seller(seller, force=force)
            results.append({'seller': seller.business_name, **row})
        except Exception as exc:
            results.append({'seller': seller.business_name, 'error': str(exc)})
    return results
